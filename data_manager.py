"""
データ処理・管理モジュール
"""
import pandas as pd
import base64
import io
from openpyxl import load_workbook
from datetime import datetime
from config import EXCEL_STRUCTURE, STAGE_MAPPING, INTEGRATED_STAGES

class DataManager:
    """データ管理クラス"""
    def __init__(self):
        self.data = None
        self.last_update = None
        self.excel_filename = None
        
    def update_data(self, contents, filename):
        """Excelデータを更新"""
        try:
            data, message = process_excel_data(contents)
            if data:
                self.data = data
                self.last_update = datetime.now()
                self.excel_filename = filename
                return True, message
            return False, message
        except Exception as e:
            return False, f"エラー: {str(e)}"
    
    def get_data(self):
        """現在のデータを取得"""
        return self.data
    
    def get_last_update(self):
        """最終更新時刻を取得"""
        if self.last_update:
            return self.last_update.strftime("%Y/%m/%d %H:%M")
        return "未更新"

# グローバルデータマネージャーインスタンス
data_manager = DataManager()

# ヘルパー関数
def format_number(value):
    """数値フォーマット"""
    if pd.isna(value) or value is None:
        return 0
    return int(round(float(value)))

def format_percentage(value):
    """パーセンテージフォーマット"""
    if pd.isna(value) or value is None:
        return 0.0
    return round(float(value), 1)

def get_column_name(col_index):
    """列インデックスから列名を取得"""
    if col_index < 26:
        return chr(65 + col_index)
    else:
        first = (col_index // 26) - 1
        second = col_index % 26
        return chr(65 + first) + chr(65 + second)

def clean_channel_names(channels):
    """チャネル名のクリーニング"""
    cleaned = []
    for ch in channels:
        if ch and str(ch).strip():
            ch_str = str(ch).strip()
            if ch_str not in ['売上高', '獲得件数（累月）', '客単価', '継続率', '指標', '合計', 'total', 'Total']:
                if ch_str in ['既存', '既存（25年以前）']:
                    ch_str = '既存'
                elif ch_str in ['フロー①', 'フロー②', 'フロー③', 'フロー④', 'フロー⑤', 'フロー']:
                    ch_str = 'フロー'
                elif ch_str == '新規（WEB）':
                    ch_str = '新規web'
                elif ch_str == '新規（法人）':
                    ch_str = '新規法人'
                elif ch_str == '新規（代理店）':
                    ch_str = '新規代理店'
                cleaned.append(ch_str)
    return list(set(cleaned))

def clean_plan_names(plans):
    """プラン名のクリーニング"""
    cleaned = []
    for plan in plans:
        if plan and str(plan).strip():
            plan_str = str(plan).strip()
            if plan_str not in ['計', '合計', 'total', 'Total', 'チャネル別小計']:
                cleaned.append(plan_str)
    return list(set(cleaned))

def filter_detail_rows(df):
    """詳細行のフィルタリング"""
    if df is None or df.empty:
        return df
    
    filtered_df = df.copy()
    
    if 'channel' in filtered_df.columns:
        exclude_channels = [
            '売上高', '獲得件数（累月）', '客単価', '継続率', '指標', 
            '合計', 'total', 'Total'
        ]
        # より柔軟なフィルタリング：明らかなヘッダー行のみを除外
        filtered_df = filtered_df[
            ~filtered_df['channel'].isin(exclude_channels)
        ]
    
    if 'plan' in filtered_df.columns:
        exclude_plans = ['計', '合計', 'total', 'Total']
        # より柔軟なフィルタリング：明らかな集計行のみを除外
        filtered_df = filtered_df[
            ~filtered_df['plan'].isin(exclude_plans)
        ]
    
    return filtered_df

def apply_filters(df, channel_filter, plan_filter):
    """フィルタ適用（階層構造対応）"""
    if df is None or df.empty:
        return df
    
    filtered_df = filter_detail_rows(df)
    
    if channel_filter:
        # チャネル名の正規化
        normalized_channels = []
        for ch in channel_filter:
            if ch == '既存':
                normalized_channels.extend(['既存', '既存（25年以前）'])
            elif ch == 'フロー':
                normalized_channels.extend(['フロー①', 'フロー②', 'フロー③', 'フロー④', 'フロー⑤', 'フロー'])
            elif ch == '新規web':
                normalized_channels.append('新規（WEB）')
            elif ch == '新規法人':
                normalized_channels.append('新規（法人）')
            elif ch == '新規代理店':
                normalized_channels.append('新規（代理店）')
            else:
                normalized_channels.append(ch)
        
        # 階層構造を考慮したフィルタリング
        selected_rows = []
        current_channel = None
        
        for idx, row in filtered_df.iterrows():
            # チャネル名が設定されている行
            if row['channel'] and str(row['channel']).strip():
                current_channel = str(row['channel']).strip()
                if current_channel in normalized_channels:
                    selected_rows.append(idx)
            # チャネル名が空の行（サブプラン行）
            elif current_channel and current_channel in normalized_channels:
                selected_rows.append(idx)
        
        filtered_df = filtered_df.loc[selected_rows] if selected_rows else filtered_df.iloc[0:0]
    
    if plan_filter:
        filtered_df = filtered_df[filtered_df['plan'].isin(plan_filter)]
    
    return filtered_df

def get_dataframe_from_store(data, section, data_type):
    """データストアからDataFrameを取得"""
    if (data and section in data and data_type in data[section] and 
        data[section][data_type] is not None):
        return pd.DataFrame(data[section][data_type])
    return None

def get_last_data_month(df, month_cols):
    """最後のデータ月を取得"""
    if df is None or df.empty or not month_cols:
        return None
    
    # 基本的な除外のみ適用（過度のフィルタリングを避ける）
    working_df = df.copy()
    if 'channel' in working_df.columns:
        # 明らかにデータ行でないもののみ除外
        exclude_channels = ['売上高', '獲得件数（累月）', '客単価', '継続率', '指標']
        working_df = working_df[~working_df['channel'].isin(exclude_channels)]
    
    if working_df.empty:
        return None
    
    for month in reversed(month_cols):
        if month in working_df.columns:
            month_total = working_df[month].sum()
            if month_total > 0:
                return month
    return None

def should_display_actual_data(df, month_cols, target_month):
    """実績データを表示すべきか判定"""
    last_data_month = get_last_data_month(df, month_cols)
    if not last_data_month or not target_month:
        return False
    
    try:
        last_idx = month_cols.index(last_data_month)
        target_idx = month_cols.index(target_month)
        return target_idx <= last_idx
    except (ValueError, IndexError):
        return False

def calculate_single_month(values, month_cols):
    """単月値の計算"""
    if len(values) <= 1:
        return values
    
    single_month_data = [values[0]]
    for i in range(1, len(values)):
        if values[i] == 0:
            single_month_data.append(0)
        else:
            single_val = values[i] - values[i-1]
            single_month_data.append(max(0, single_val))
    
    return single_month_data

def calculate_cumulative(values):
    """累月値の計算"""
    if len(values) <= 1:
        return values
    
    cumulative_data = [values[0]]
    for i in range(1, len(values)):
        cumulative_val = cumulative_data[i-1] + values[i]
        cumulative_data.append(cumulative_val)
    
    return cumulative_data

def process_excel_data(contents):
    """Excelファイルを処理してPDCAデータを抽出"""
    try:
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)
        
        wb = load_workbook(io.BytesIO(decoded), data_only=True)
        
        if EXCEL_STRUCTURE['sheet_name'] not in wb.sheetnames:
            return None, f"「{EXCEL_STRUCTURE['sheet_name']}」シートが見つかりません"
        
        ws = wb[EXCEL_STRUCTURE['sheet_name']]
        
        data = {}
        
        for section_name, (start_row, end_row) in EXCEL_STRUCTURE['sections'].items():
            section_data = {}
            
            for data_type, (start_col, end_col) in EXCEL_STRUCTURE['col_ranges'].items():
                type_data = []
                
                # 月情報の取得
                months = []
                for col in range(start_col, end_col + 1):
                    col_name = get_column_name(col)
                    cell = ws[f'{col_name}4']
                    if cell.value is not None:
                        if isinstance(cell.value, (int, float)) and 1 <= cell.value <= 12:
                            months.append(f"{int(cell.value)}月")
                        elif cell.value == '合計':
                            months.append('合計')
                
                # データの取得
                for row in range(start_row, end_row + 1):
                    channel = ws[f'A{row}'].value if ws[f'A{row}'].value else ""
                    plan = ws[f'B{row}'].value if ws[f'B{row}'].value else ""
                    
                    if not channel and not plan:
                        continue
                    
                    row_data = {
                        'channel': channel,
                        'plan': plan,
                        'section': section_name
                    }
                    
                    for i, col in enumerate(range(start_col, end_col + 1)):
                        col_name = get_column_name(col)
                        cell = ws[f'{col_name}{row}']
                        value = cell.value if cell.value is not None else 0
                        
                        if i < len(months):
                            row_data[months[i]] = value
                    
                    type_data.append(row_data)
                
                section_data[data_type] = type_data
            
            data[section_name] = section_data
        
        # データの数値変換
        processed_data = {}
        for section, section_data in data.items():
            processed_data[section] = {}
            for data_type, type_data in section_data.items():
                if type_data:
                    df = pd.DataFrame(type_data)
                    month_cols = [col for col in df.columns if col.endswith('月') or col == '合計']
                    for col in month_cols:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    processed_data[section][data_type] = df.to_dict('records')
        
        return processed_data, "データの読み込みが完了しました"
        
    except Exception as e:
        return None, f"ファイル処理エラー: {str(e)}"

def calculate_kpi_values(data, section, selected_month, data_type, period_type, channel_filter, plan_filter):
    """KPI値の計算"""
    actual_df = get_dataframe_from_store(data, section, 'actual')
    budget_df = get_dataframe_from_store(data, section, 'budget')
    
    actual_value = 0
    budget_value = 0
    
    if actual_df is not None and not actual_df.empty:
        month_cols = [col for col in actual_df.columns if col.endswith('月')]
        if should_display_actual_data(actual_df, month_cols, selected_month):
            filtered_df = apply_filters(actual_df, channel_filter, plan_filter)
            if not filtered_df.empty and selected_month in filtered_df.columns:
                # 売上高の単月計算: データは既に単月形式なので直接使用
                if period_type == 'single' and section == 'sales':
                    actual_value = filtered_df[selected_month].sum()
                # 獲得件数の単月計算: 累月データから前月を差し引く
                elif period_type == 'single' and section == 'acquisition':
                    month_idx = month_cols.index(selected_month)
                    current_val = filtered_df[selected_month].sum()
                    if month_idx > 0:
                        prev_month = month_cols[month_idx - 1]
                        prev_val = filtered_df[prev_month].sum() if prev_month in filtered_df.columns else 0
                        actual_value = max(0, current_val - prev_val)
                    else:
                        actual_value = current_val
                # 売上高の累月計算: 単月データから累計を計算
                elif period_type == 'cumulative' and section == 'sales':
                    month_idx = month_cols.index(selected_month)
                    cumulative_sum = 0
                    for i in range(month_idx + 1):
                        month = month_cols[i]
                        if month in filtered_df.columns:
                            cumulative_sum += filtered_df[month].sum()
                    actual_value = cumulative_sum
                # その他のケース: 直接値を使用
                else:
                    actual_value = filtered_df[selected_month].sum()
    
    if budget_df is not None and not budget_df.empty:
        filtered_budget_df = apply_filters(budget_df, channel_filter, plan_filter)
        if not filtered_budget_df.empty and selected_month in filtered_budget_df.columns:
            month_cols = [col for col in budget_df.columns if col.endswith('月')]
            # 売上高の単月計算: データは既に単月形式なので直接使用
            if period_type == 'single' and section == 'sales':
                budget_value = filtered_budget_df[selected_month].sum()
            # 獲得件数の単月計算: 累月データから前月を差し引く
            elif period_type == 'single' and section == 'acquisition':
                month_idx = month_cols.index(selected_month)
                current_val = filtered_budget_df[selected_month].sum()
                if month_idx > 0:
                    prev_month = month_cols[month_idx - 1]
                    prev_val = filtered_budget_df[prev_month].sum() if prev_month in filtered_budget_df.columns else 0
                    budget_value = max(0, current_val - prev_val)
                else:
                    budget_value = current_val
            # 売上高の累月計算: 単月データから累計を計算
            elif period_type == 'cumulative' and section == 'sales':
                month_idx = month_cols.index(selected_month)
                cumulative_sum = 0
                for i in range(month_idx + 1):
                    month = month_cols[i]
                    if month in filtered_budget_df.columns:
                        cumulative_sum += filtered_budget_df[month].sum()
                budget_value = cumulative_sum
            # その他のケース: 直接値を使用
            else:
                budget_value = filtered_budget_df[selected_month].sum()
    
    return actual_value, budget_value

def get_monthly_trend_data(data, section, data_type, period_type, channel_filter=None, plan_filter=None, target_item=None, stage_name=None):
    """月別トレンドデータを取得（スパークライン用）"""
    actual_df = get_dataframe_from_store(data, section, 'actual')
    budget_df = get_dataframe_from_store(data, section, 'budget')
    
    actual_values = []
    budget_values = []
    months = []
    last_data_month_index = -1
    
    if actual_df is not None and not actual_df.empty:
        month_cols = [col for col in actual_df.columns if col.endswith('月')]
        
        # 指標セクションの場合、ステージベースでフィルター
        if section == 'indicators' and stage_name:
            from config import STAGE_MAPPING
            if stage_name in STAGE_MAPPING:
                stage_rows = STAGE_MAPPING[stage_name]
                filtered_df = pd.DataFrame()
                
                for row_idx in stage_rows:
                    df_idx = row_idx - 101
                    if 0 <= df_idx < len(actual_df):
                        filtered_df = pd.concat([filtered_df, actual_df.iloc[[df_idx]]])
            else:
                filtered_df = actual_df.copy()
        else:
            # 特定のアイテム（チャネルやプラン）でフィルター
            filtered_df = actual_df.copy()
            
            # プラン指定がある場合は、チャネルフィルターを無視する
            if target_item and 'plan' in filtered_df.columns and plan_filter:
                # プランでフィルター（チャネルフィルターは適用しない）
                filtered_df = apply_filters(filtered_df, None, plan_filter)
            elif target_item and 'channel' in filtered_df.columns:
                # チャネルでフィルター - 正規化してマッチング
                # チャネル名の逆マッピング（表示名→内部名）
                channel_mapping = {
                    '新規web': '新規（WEB）',
                    '新規法人': '新規（法人）',
                    '新規代理店': '新規（代理店）',
                    'フロー': ['フロー①', 'フロー②', 'フロー③', 'フロー④', 'フロー⑤', 'フロー'],
                    '既存': ['既存', '既存（25年以前）']
                }
                
                if target_item in channel_mapping:
                    internal_names = channel_mapping[target_item]
                    if isinstance(internal_names, list):
                        filtered_df = filtered_df[filtered_df['channel'].isin(internal_names)]
                    else:
                        filtered_df = filtered_df[filtered_df['channel'] == internal_names]
                else:
                    filtered_df = filtered_df[filtered_df['channel'] == target_item]
                # target_itemが指定されている場合は、追加のチャネルフィルターは適用しない
                filtered_df = apply_filters(filtered_df, None, plan_filter)
            else:
                # 通常のフィルター適用
                filtered_df = apply_filters(filtered_df, channel_filter, plan_filter)
        
        # 実データのある最後の月を見つける（0でない値がある月）
        for i, month in enumerate(month_cols):
            if month in filtered_df.columns:
                month_sum = filtered_df[month].sum()
                if should_display_actual_data(filtered_df, month_cols, month) and month_sum > 0:
                    last_data_month_index = i
        
        # 実データのある月まででデータを切り取る
        for i, month in enumerate(month_cols):
            if i > last_data_month_index and last_data_month_index >= 0:
                break  # 実データの終了後は配列に追加しない
                
            if month in filtered_df.columns:
                months.append(month)
                if should_display_actual_data(filtered_df, month_cols, month):
                    # 獲得件数の単月計算: 累月データから前月を差し引く
                    if period_type == 'single' and section == 'acquisition':
                        current_val = filtered_df[month].sum()
                        if i > 0:
                            prev_month = month_cols[i - 1]
                            prev_val = filtered_df[prev_month].sum() if prev_month in filtered_df.columns else 0
                            single_val = max(0, current_val - prev_val)
                            actual_values.append(single_val)
                        else:
                            actual_values.append(current_val)
                    else:
                        actual_values.append(filtered_df[month].sum())
                else:
                    actual_values.append(0)
    
    if budget_df is not None and not budget_df.empty:
        # 計画データも同様に処理（実データと同じ長さに合わせる）
        if section == 'indicators' and stage_name:
            from config import STAGE_MAPPING
            if stage_name in STAGE_MAPPING:
                stage_rows = STAGE_MAPPING[stage_name]
                filtered_budget_df = pd.DataFrame()
                
                for row_idx in stage_rows:
                    df_idx = row_idx - 101
                    if 0 <= df_idx < len(budget_df):
                        filtered_budget_df = pd.concat([filtered_budget_df, budget_df.iloc[[df_idx]]])
            else:
                filtered_budget_df = budget_df.copy()
        else:
            filtered_budget_df = budget_df.copy()
            
            # プラン指定がある場合は、チャネルフィルターを無視する
            if target_item and 'plan' in filtered_budget_df.columns and plan_filter:
                # プランでフィルター（チャネルフィルターは適用しない）
                filtered_budget_df = apply_filters(filtered_budget_df, None, plan_filter)
            elif target_item and 'channel' in filtered_budget_df.columns:
                # チャネルでフィルター - 正規化してマッチング
                # チャネル名の逆マッピング（表示名→内部名）
                channel_mapping = {
                    '新規web': '新規（WEB）',
                    '新規法人': '新規（法人）',
                    '新規代理店': '新規（代理店）',
                    'フロー': ['フロー①', 'フロー②', 'フロー③', 'フロー④', 'フロー⑤', 'フロー'],
                    '既存': ['既存', '既存（25年以前）']
                }
                
                if target_item in channel_mapping:
                    internal_names = channel_mapping[target_item]
                    if isinstance(internal_names, list):
                        filtered_budget_df = filtered_budget_df[filtered_budget_df['channel'].isin(internal_names)]
                    else:
                        filtered_budget_df = filtered_budget_df[filtered_budget_df['channel'] == internal_names]
                else:
                    filtered_budget_df = filtered_budget_df[filtered_budget_df['channel'] == target_item]
                
                # target_itemが指定されている場合は、追加のチャネルフィルターは適用しない
                filtered_budget_df = apply_filters(filtered_budget_df, None, plan_filter)
            else:
                # 通常のフィルター適用
                filtered_budget_df = apply_filters(filtered_budget_df, channel_filter, plan_filter)
        
        # 計画値は12月まですべて取得
        for i, month in enumerate(month_cols):
            if month in filtered_budget_df.columns:
                # 獲得件数の単月計算: 累月データから前月を差し引く
                if period_type == 'single' and section == 'acquisition':
                    current_val = filtered_budget_df[month].sum()
                    if i > 0:
                        prev_month = month_cols[i - 1]
                        prev_val = filtered_budget_df[prev_month].sum() if prev_month in filtered_budget_df.columns else 0
                        budget_values.append(max(0, current_val - prev_val))
                    else:
                        budget_values.append(current_val)
                else:
                    budget_values.append(filtered_budget_df[month].sum())
            else:
                budget_values.append(0)
    
    # 実際にデータがある月数を記録
    actual_months_count = len(actual_values)
    
    # actual_valuesを12月まで拡張（残りは0で埋める）
    while len(actual_values) < len(budget_values):
        actual_values.append(0)
    
    return {
        'months': month_cols,  # 12月まですべての月を返す
        'actual_values': actual_values,
        'budget_values': budget_values,
        'actual_months': actual_months_count  # 実際のデータがある月数
    }